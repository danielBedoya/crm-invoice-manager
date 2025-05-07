import openpyxl
import csv
import io

from dashboard.constants import REQUIRED_COLUMNS


def process_excel_file(uploaded_file):
    """
    Processes an uploaded Excel file and validates its structure.

    This function checks if the uploaded Excel file contains the required columns
    and extracts its data into a list of dictionaries. If any required columns
    are missing, it returns an error message.

    Args:
        uploaded_file (File): The uploaded Excel file to be processed.

    Returns:
        tuple: A tuple containing:
            - data (list): A list of dictionaries representing the rows in the file.
              Each dictionary maps column headers to their respective values.
            - error (str or None): An error message if validation fails, or None if successful.
    """

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        return None, f"Faltan las siguientes columnas requeridas: {', '.join(missing)}"

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))

    return data, None


def process_csv_file(uploaded_file):
    """
    Processes an uploaded CSV file and validates its structure.

    This function checks if the uploaded CSV file contains the required columns
    and extracts its data into a list of dictionaries. If any required columns
    are missing, it returns an error message.

    Args:
        uploaded_file (File): The uploaded CSV file to be processed.

    Returns:
        tuple: A tuple containing:
            - data (list): A list of dictionaries representing the rows in the file.
              Each dictionary maps column headers to their respective values.
            - error (str or None): An error message if validation fails, or None if successful.
    """

    decoded_file = uploaded_file.read().decode("utf-8")
    io_string = io.StringIO(decoded_file)
    reader = csv.DictReader(io_string)

    headers = reader.fieldnames
    if not headers:
        return None, "El archivo CSV no contiene encabezados."

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        return None, f"Faltan las siguientes columnas requeridas: {', '.join(missing)}"

    data = [row for row in reader]
    return data, None
